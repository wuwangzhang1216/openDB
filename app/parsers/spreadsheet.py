"""Spreadsheet parser for XLSX and CSV.

XLSX: openpyxl, one page per sheet (split at 100 rows).
CSV: pandas, treated as single-sheet XLSX.
"""

from __future__ import annotations

import logging
from datetime import datetime, date, time
from pathlib import Path

from app.parsers.base import Page, ParseResult
from app.parsers.registry import register

logger = logging.getLogger(__name__)

ROWS_PER_PAGE = 100
HEADER_REPEAT_INTERVAL = 50


def _serialize_cell(value):
    """Serialize a cell value to a JSON-safe type."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.isoformat()
    if isinstance(value, (int, float)):
        return value
    if isinstance(value, bool):
        return value
    return str(value)


class XlsxParser:
    def parse(self, file_path: Path) -> ParseResult:
        import openpyxl

        wb = openpyxl.load_workbook(str(file_path), read_only=True, data_only=True)
        try:
            pages: list[Page] = []
            page_num = 0

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = list(ws.iter_rows(values_only=True))
                if not rows:
                    page_num += 1
                    pages.append(
                        Page(
                            page_number=page_num,
                            section_title=sheet_name,
                            text=f"(empty sheet: {sheet_name})",
                        )
                    )
                    continue

                # First row as headers
                headers = [str(c) if c is not None else "" for c in rows[0]]
                data_rows = rows[1:]

                # Split into chunks of ROWS_PER_PAGE
                for chunk_start in range(0, max(len(data_rows), 1), ROWS_PER_PAGE):
                    page_num += 1
                    chunk = data_rows[chunk_start: chunk_start + ROWS_PER_PAGE]

                    text = self._format_rows(headers, chunk, chunk_start)
                    suffix = ""
                    if len(data_rows) > ROWS_PER_PAGE:
                        row_start = chunk_start + 2  # +2: 1-indexed + header row
                        row_end = row_start + len(chunk) - 1
                        suffix = f" - rows {row_start}-{row_end}"

                    pages.append(
                        Page(
                            page_number=page_num,
                            section_title=f"{sheet_name}{suffix}" if suffix else sheet_name,
                            text=text.strip(),
                        )
                    )

            metadata = self._extract_metadata(wb)
        finally:
            wb.close()
        return ParseResult(pages=pages, extracted_metadata=metadata)

    def _format_rows(
        self, headers: list[str], rows: list[tuple], chunk_start: int
    ) -> str:
        """Format rows as 'Col: val | Col: val' with periodic header repeats."""
        lines: list[str] = []

        # Column info header
        col_info = f"Columns: {' | '.join(headers)}"
        lines.append(col_info)
        lines.append("")

        for i, row in enumerate(rows):
            # Repeat headers every HEADER_REPEAT_INTERVAL rows
            if i > 0 and i % HEADER_REPEAT_INTERVAL == 0:
                lines.append("")
                lines.append(col_info)
                lines.append("")

            cells = [str(c) if c is not None else "" for c in row]
            pairs = [f"{h}: {v}" for h, v in zip(headers, cells)]
            lines.append(" | ".join(pairs))

        return "\n".join(lines)

    def parse_structured(
        self, file_path: Path, sheet_filter: list[str] | None = None,
    ) -> dict:
        """Return structured JSON with columns and rows per sheet."""
        import openpyxl

        wb = openpyxl.load_workbook(str(file_path), read_only=True, data_only=True)
        try:
            sheets = []
            for sheet_name in wb.sheetnames:
                if sheet_filter and sheet_name not in sheet_filter:
                    continue
                ws = wb[sheet_name]
                rows = list(ws.iter_rows(values_only=True))
                if not rows:
                    sheets.append({
                        "name": sheet_name,
                        "columns": [],
                        "rows": [],
                        "total_rows": 0,
                    })
                    continue
                columns = [str(c) if c is not None else "" for c in rows[0]]
                data_rows = [
                    [_serialize_cell(c) for c in row] for row in rows[1:]
                ]
                sheets.append({
                    "name": sheet_name,
                    "columns": columns,
                    "rows": data_rows,
                    "total_rows": len(data_rows),
                })
            return {"sheets": sheets}
        finally:
            wb.close()

    def _extract_metadata(self, wb) -> dict:
        """Extract XLSX workbook properties."""
        result = {}
        try:
            props = wb.properties
            if props.creator:
                result["author"] = props.creator
            if props.title:
                result["title"] = props.title
            if hasattr(props, "company") and props.company:
                result["company"] = props.company
            if props.created:
                result["created"] = props.created.isoformat()
            if props.modified:
                result["modified"] = props.modified.isoformat()
            result["sheet_names"] = wb.sheetnames
        except Exception:
            logger.warning("XLSX metadata extraction failed", exc_info=True)
        return result


class CsvParser:
    def parse(self, file_path: Path) -> ParseResult:
        import pandas as pd

        try:
            df = pd.read_csv(str(file_path))
        except Exception:
            # Fallback encodings
            for enc in ("latin-1", "utf-16", "cp1252"):
                try:
                    df = pd.read_csv(str(file_path), encoding=enc)
                    break
                except Exception:
                    continue
            else:
                return ParseResult(
                    pages=[Page(page_number=1, section_title=None, text="(unreadable CSV)")]
                )

        headers = [str(c) for c in df.columns]
        rows = [tuple(row) for _, row in df.iterrows()]

        pages: list[Page] = []
        page_num = 0

        for chunk_start in range(0, max(len(rows), 1), ROWS_PER_PAGE):
            page_num += 1
            chunk = rows[chunk_start: chunk_start + ROWS_PER_PAGE]
            text = self._format_rows(headers, chunk, chunk_start)

            suffix = ""
            if len(rows) > ROWS_PER_PAGE:
                row_start = chunk_start + 2
                row_end = row_start + len(chunk) - 1
                suffix = f" - rows {row_start}-{row_end}"

            pages.append(
                Page(
                    page_number=page_num,
                    section_title=f"Data{suffix}" if suffix else None,
                    text=text.strip(),
                )
            )

        return ParseResult(pages=pages)

    def parse_structured(
        self, file_path: Path, sheet_filter: list[str] | None = None,
    ) -> dict:
        """Return structured JSON with columns and rows."""
        import pandas as pd

        try:
            df = pd.read_csv(str(file_path))
        except Exception:
            for enc in ("latin-1", "utf-16", "cp1252"):
                try:
                    df = pd.read_csv(str(file_path), encoding=enc)
                    break
                except Exception:
                    continue
            else:
                return {"sheets": [{"name": "Data", "columns": [], "rows": [], "total_rows": 0}]}

        columns = [str(c) for c in df.columns]
        # Replace NaN with None for JSON serialization
        rows = df.where(df.notna(), None).values.tolist()
        # Serialize each cell
        rows = [[_serialize_cell(c) for c in row] for row in rows]

        return {
            "sheets": [{
                "name": "Data",
                "columns": columns,
                "rows": rows,
                "total_rows": len(rows),
            }]
        }

    def _format_rows(
        self, headers: list[str], rows: list[tuple], chunk_start: int
    ) -> str:
        """Format rows as 'Col: val | Col: val'."""
        lines: list[str] = []
        col_info = f"Columns: {' | '.join(headers)}"
        lines.append(col_info)
        lines.append("")

        for i, row in enumerate(rows):
            if i > 0 and i % HEADER_REPEAT_INTERVAL == 0:
                lines.append("")
                lines.append(col_info)
                lines.append("")

            cells = [str(c) if c is not None else "" for c in row]
            pairs = [f"{h}: {v}" for h, v in zip(headers, cells)]
            lines.append(" | ".join(pairs))

        return "\n".join(lines)


register(
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    XlsxParser(),
)
register("application/vnd.ms-excel", XlsxParser())
register("text/csv", CsvParser())
