"""Unit tests for structured spreadsheet output (Gap 2)."""

from pathlib import Path

import pytest

from opendb_core.parsers.spreadsheet import XlsxParser, CsvParser, _serialize_cell


# ---------------------------------------------------------------------------
# _serialize_cell
# ---------------------------------------------------------------------------

class TestSerializeCell:
    def test_none(self) -> None:
        assert _serialize_cell(None) is None

    def test_int(self) -> None:
        assert _serialize_cell(42) == 42

    def test_float(self) -> None:
        assert _serialize_cell(3.14) == 3.14

    def test_bool(self) -> None:
        assert _serialize_cell(True) is True

    def test_string(self) -> None:
        assert _serialize_cell("hello") == "hello"

    def test_datetime(self) -> None:
        from datetime import datetime
        dt = datetime(2024, 3, 15, 10, 30, 0)
        assert _serialize_cell(dt) == "2024-03-15T10:30:00"

    def test_date(self) -> None:
        from datetime import date
        d = date(2024, 3, 15)
        assert _serialize_cell(d) == "2024-03-15"

    def test_time(self) -> None:
        from datetime import time
        t = time(10, 30, 0)
        assert _serialize_cell(t) == "10:30:00"

    def test_other_type_becomes_string(self) -> None:
        assert _serialize_cell([1, 2]) == "[1, 2]"


# ---------------------------------------------------------------------------
# XlsxParser.parse_structured
# ---------------------------------------------------------------------------

class TestXlsxParseStructured:
    @pytest.fixture
    def sample_xlsx(self, tmp_path: Path) -> Path:
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Revenue"
        ws.append(["Month", "Amount", "Category"])
        ws.append(["Jan", 1000, "Sales"])
        ws.append(["Feb", 1500, "Sales"])
        ws.append(["Mar", 2000, "Consulting"])

        ws2 = wb.create_sheet("Expenses")
        ws2.append(["Item", "Cost"])
        ws2.append(["Rent", 3000])
        ws2.append(["Utils", 500])

        wb.create_sheet("Empty")  # empty sheet

        path = tmp_path / "finance.xlsx"
        wb.save(path)
        wb.close()
        return path

    def test_basic_structured_output(self, sample_xlsx: Path) -> None:
        parser = XlsxParser()
        result = parser.parse_structured(sample_xlsx)

        assert "sheets" in result
        assert len(result["sheets"]) == 3

        revenue = result["sheets"][0]
        assert revenue["name"] == "Revenue"
        assert revenue["columns"] == ["Month", "Amount", "Category"]
        assert revenue["total_rows"] == 3
        assert revenue["rows"][0] == ["Jan", 1000, "Sales"]
        assert revenue["rows"][2] == ["Mar", 2000, "Consulting"]

    def test_empty_sheet(self, sample_xlsx: Path) -> None:
        parser = XlsxParser()
        result = parser.parse_structured(sample_xlsx)

        empty = result["sheets"][2]
        assert empty["name"] == "Empty"
        assert empty["columns"] == []
        assert empty["rows"] == []
        assert empty["total_rows"] == 0

    def test_sheet_filter(self, sample_xlsx: Path) -> None:
        parser = XlsxParser()
        result = parser.parse_structured(sample_xlsx, sheet_filter=["Expenses"])

        assert len(result["sheets"]) == 1
        assert result["sheets"][0]["name"] == "Expenses"
        assert result["sheets"][0]["columns"] == ["Item", "Cost"]
        assert result["sheets"][0]["total_rows"] == 2

    def test_sheet_filter_no_match(self, sample_xlsx: Path) -> None:
        parser = XlsxParser()
        result = parser.parse_structured(sample_xlsx, sheet_filter=["NonExistent"])
        assert result["sheets"] == []

    def test_no_filter_returns_all(self, sample_xlsx: Path) -> None:
        parser = XlsxParser()
        result = parser.parse_structured(sample_xlsx, sheet_filter=None)
        assert len(result["sheets"]) == 3


# ---------------------------------------------------------------------------
# CsvParser.parse_structured
# ---------------------------------------------------------------------------

FIXTURES = Path(__file__).parent / "fixtures"


class TestCsvParseStructured:
    def test_basic_csv(self) -> None:
        parser = CsvParser()
        result = parser.parse_structured(FIXTURES / "sample.csv")

        assert "sheets" in result
        assert len(result["sheets"]) == 1
        sheet = result["sheets"][0]
        assert sheet["name"] == "Data"
        assert len(sheet["columns"]) > 0
        assert sheet["total_rows"] > 0

    def test_simple_csv(self, tmp_path: Path) -> None:
        f = tmp_path / "test.csv"
        f.write_text("name,age,city\nAlice,30,NYC\nBob,25,LA\n")

        parser = CsvParser()
        result = parser.parse_structured(f)

        sheet = result["sheets"][0]
        assert sheet["columns"] == ["name", "age", "city"]
        assert sheet["total_rows"] == 2
        assert sheet["rows"][0][0] == "Alice"
        assert sheet["rows"][1][0] == "Bob"

    def test_empty_csv(self, tmp_path: Path) -> None:
        f = tmp_path / "empty.csv"
        f.write_text("")

        parser = CsvParser()
        result = parser.parse_structured(f)

        sheet = result["sheets"][0]
        assert sheet["total_rows"] == 0
        assert sheet["columns"] == []

    def test_sheet_filter_ignored_for_csv(self, tmp_path: Path) -> None:
        """CSV only has one 'Data' sheet; filter param is accepted but benign."""
        f = tmp_path / "test.csv"
        f.write_text("a,b\n1,2\n")

        parser = CsvParser()
        result = parser.parse_structured(f, sheet_filter=["Data"])
        assert len(result["sheets"]) == 1
